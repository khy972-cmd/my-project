from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib import parse, request
from urllib.error import HTTPError

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SITE_XLSX = Path(r"c:\Users\bobo\Desktop\site-manager-main_260302\INOPNC_FN-main\site.xlsx")
DEFAULT_WORKERS_CSV = Path(r"c:\Users\bobo\Desktop\site-manager-main_260302\INOPNC_FN-main\workers_rows.csv")
DEFAULT_WORKER_AFFILIATION = "\uC774\uB178\uD53C\uC564\uC528"
SPECIAL_ROLE_TARGETS = {
    "김재형": "admin",
    "김혜영": "admin",
    "송용호": "manager",
    "권용호": "manager",
}
SITE_STATUS_MAP = {
    "active": "진행중",
    "scheduled": "예정",
    "completed": "완료",
}


@dataclass
class SupabaseClient:
    base_url: str
    api_key: str
    access_token: str

    @property
    def headers(self) -> dict[str, str]:
        return {
            "apikey": self.api_key,
            "Authorization": f"Bearer {self.access_token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }

    def rest(self, path: str, method: str = "GET", payload: Any | None = None, extra_headers: dict[str, str] | None = None):
        url = f"{self.base_url}/rest/v1/{path}"
        headers = self.headers.copy()
        if extra_headers:
            headers.update(extra_headers)
        data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = request.Request(url, data=data, method=method, headers=headers)
        with request.urlopen(req) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else None


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    for line in (ROOT / ".env.local").read_text(encoding="utf-8").splitlines():
        if not line or line.strip().startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def auth_client(email: str, password: str) -> SupabaseClient:
    env = load_env()
    base_url = env["VITE_SUPABASE_URL"].rstrip("/")
    api_key = env["VITE_SUPABASE_ANON_KEY"]
    req = request.Request(
        f"{base_url}/auth/v1/token?grant_type=password",
        data=json.dumps({"email": email, "password": password}).encode("utf-8"),
        headers={"apikey": api_key, "Content-Type": "application/json"},
        method="POST",
    )
    with request.urlopen(req) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return SupabaseClient(base_url=base_url, api_key=api_key, access_token=payload["access_token"])


def read_site_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data = [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(value is not None and str(value).strip() != "" for value in row)
    ]
    return data


def read_worker_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [row for row in reader if any((value or "").strip() for value in row.values())]


def fetch_rows(client: SupabaseClient, table: str, select: str, order: str | None = None, limit: int | None = None):
    params = {"select": select}
    if order:
        params["order"] = order
    if limit is not None:
        params["limit"] = str(limit)
    query = parse.urlencode(params, safe="(),*")
    return client.rest(f"{table}?{query}")


def normalize_status(raw: Any) -> str:
    return SITE_STATUS_MAP.get(str(raw or "").strip().lower(), str(raw or "").strip() or "예정")


def build_site_plan(site_rows: list[dict[str, Any]], db_sites: list[dict[str, Any]]) -> dict[str, Any]:
    db_by_id = {row["id"]: row for row in db_sites}
    inserts = 0
    updates = 0
    seen_names: dict[str, str] = {}
    duplicate_names: list[dict[str, str]] = []

    for row in site_rows:
        site_id = str(row.get("id") or "").strip()
        site_name = str(row.get("name") or "").strip()
        normalized = normalize_status(row.get("status"))
        existing = db_by_id.get(site_id)
        if existing is None:
            inserts += 1
        else:
            if (
                str(existing.get("name") or "") != site_name
                or str(existing.get("status") or "") != normalized
                or str(existing.get("builder") or "") != str(row.get("builder") or "")
                or str(existing.get("company_name") or "") != str(row.get("company_name") or "")
                or str(existing.get("source_dataset") or "") != "site.xlsx"
            ):
                updates += 1

        if site_name in seen_names:
            duplicate_names.append({
                "name": site_name,
                "first_id": seen_names[site_name],
                "second_id": site_id,
            })
        else:
            seen_names[site_name] = site_id

    return {
        "insert_count": inserts,
        "update_count": updates,
        "duplicate_name_collisions": duplicate_names,
        "db_overlap_count": len(set(db_by_id) & {str(row.get("id") or "").strip() for row in site_rows}),
    }


def build_role_plan(profiles: list[dict[str, Any]], roles: list[dict[str, Any]]) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    role_map = {row["user_id"]: row["role"] for row in roles}
    profiles_by_name: dict[str, list[dict[str, Any]]] = {}
    for row in profiles:
        name = str(row.get("name") or "").strip()
        profiles_by_name.setdefault(name, []).append(row)

    partner_ids = {row["user_id"] for row in roles if row["role"] == "partner"}
    dry_run: list[dict[str, str]] = []
    blockers: list[dict[str, Any]] = []

    for row in profiles:
        name = str(row.get("name") or "").strip()
        user_id = row["user_id"]
        current_role = role_map.get(user_id, "worker")
        if name in SPECIAL_ROLE_TARGETS:
            next_role = SPECIAL_ROLE_TARGETS[name]
            reason = f"special:{name}"
        elif current_role == "partner":
            next_role = "partner"
            reason = "existing_partner_preserved"
        else:
            next_role = "worker"
            reason = "default_worker"
        dry_run.append({
            "user_id": user_id,
            "name": name,
            "current_role": current_role,
            "next_role": next_role,
            "reason": reason,
        })

    for name, next_role in SPECIAL_ROLE_TARGETS.items():
        matches = profiles_by_name.get(name, [])
        if len(matches) == 0:
            blockers.append({"type": "missing_special_profile", "name": name, "target_role": next_role})
        elif len(matches) > 1:
            blockers.append({
                "type": "duplicate_special_profile",
                "name": name,
                "target_role": next_role,
                "user_ids": [row["user_id"] for row in matches],
            })
        elif matches[0]["user_id"] in partner_ids:
            blockers.append({
                "type": "partner_role_conflict",
                "name": name,
                "target_role": next_role,
                "user_id": matches[0]["user_id"],
            })

    return dry_run, blockers


def build_directory_rows(
    worker_rows: list[dict[str, str]],
    profiles: list[dict[str, Any]],
    roles: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    role_map = {row["user_id"]: row["role"] for row in roles}
    profiles_by_name: dict[str, list[dict[str, Any]]] = {}
    for row in profiles:
        name = str(row.get("name") or "").strip()
        profiles_by_name.setdefault(name, []).append(row)

    directory_rows: list[dict[str, Any]] = []
    unmatched: list[str] = []
    ambiguous: list[dict[str, Any]] = []
    partner_conflicts: list[dict[str, Any]] = []

    for row in worker_rows:
        name = str(row.get("name") or "").strip()
        matches = profiles_by_name.get(name, [])
        linked_user_id = None
        role = "worker"
        notes = None

        if len(matches) == 1 and role_map.get(matches[0]["user_id"]) != "partner":
            linked_user_id = matches[0]["user_id"]
        elif len(matches) == 0:
            unmatched.append(name)
        elif len(matches) > 1:
            ambiguous.append({"name": name, "user_ids": [match["user_id"] for match in matches]})
            notes = "profile_name_duplicate"
        elif len(matches) == 1 and role_map.get(matches[0]["user_id"]) == "partner":
            partner_conflicts.append({"name": name, "user_id": matches[0]["user_id"]})
            notes = "partner_conflict"

        if name in SPECIAL_ROLE_TARGETS and linked_user_id:
            role = SPECIAL_ROLE_TARGETS[name]

        directory_rows.append({
            "source_worker_id": row["id"],
            "linked_user_id": linked_user_id,
            "name": name,
            "phone": matches[0].get("phone") if len(matches) == 1 else None,
            "affiliation": DEFAULT_WORKER_AFFILIATION,
            "role": role,
            "daily": int(row["daily"]) if str(row.get("daily") or "").strip() else None,
            "notes": notes,
            "is_active": True,
            "source": "workers_rows.csv",
        })

    stats = {
        "exact_unique_matches": sum(1 for row in directory_rows if row["linked_user_id"]),
        "unmatched_count": len(unmatched),
        "ambiguous_count": len(ambiguous),
        "partner_conflict_count": len(partner_conflicts),
        "unmatched_names": unmatched,
        "ambiguous_names": ambiguous,
        "partner_conflicts": partner_conflicts,
    }
    return directory_rows, stats


def merge_pending_directory_rows(
    directory_rows: list[dict[str, Any]],
    pending_assignments: list[dict[str, Any]],
    existing_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    existing_by_name = {str(row.get("name") or "").strip(): row for row in existing_rows}
    desired_names = {str(row.get("name") or "").strip() for row in directory_rows}
    merged_rows = list(directory_rows)

    for assignment in pending_assignments:
        status = str(assignment.get("status") or "").strip().lower()
        if status == "cancelled":
            continue

        name = str(assignment.get("reserved_name") or "").strip()
        if not name or name in desired_names:
            continue

        existing = existing_by_name.get(name, {})
        notes = None if assignment.get("linked_user_id") else "auth_account_required"

        merged_rows.append({
            "source_worker_id": existing.get("source_worker_id"),
            "linked_user_id": assignment.get("linked_user_id") or existing.get("linked_user_id"),
            "name": name,
            "phone": existing.get("phone"),
            "affiliation": DEFAULT_WORKER_AFFILIATION,
            "role": str(assignment.get("reserved_role") or existing.get("role") or "worker"),
            "daily": existing.get("daily"),
            "notes": notes,
            "is_active": True,
            "source": "pending_role_assignment",
        })
        desired_names.add(name)

    return merged_rows


def build_directory_plan(directory_rows: list[dict[str, Any]], existing_rows: list[dict[str, Any]]) -> dict[str, Any]:
    desired_names = {str(row.get("name") or "").strip() for row in directory_rows}
    active_existing_rows = [row for row in existing_rows if row.get("is_active")]
    stale_rows = [
        row
        for row in active_existing_rows
        if str(row.get("name") or "").strip() not in desired_names
    ]
    stale_sources = Counter(str(row.get("source") or "unknown") for row in stale_rows)

    return {
        "upsert_count": len(directory_rows),
        "active_existing_count": len(active_existing_rows),
        "deactivate_count": len(stale_rows),
        "deactivate_ids": [row["id"] for row in stale_rows],
        "deactivate_names": sorted(str(row.get("name") or "").strip() for row in stale_rows),
        "deactivate_source_breakdown": dict(stale_sources),
    }


def apply_sites(client: SupabaseClient, site_rows: list[dict[str, Any]]) -> int:
    payload = [
        {
            "id": str(row.get("id") or "").strip(),
            "name": str(row.get("name") or "").strip(),
            "status": normalize_status(row.get("status")),
            "builder": str(row.get("builder") or "").strip() or None,
            "company_name": str(row.get("company_name") or "").strip() or None,
            "created_at": str(row.get("created_at") or "").strip() or None,
            "source_dataset": "site.xlsx",
        }
        for row in site_rows
    ]
    client.rest(
        "sites?on_conflict=id",
        method="POST",
        payload=payload,
        extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"},
    )
    return len(payload)


def apply_directory(client: SupabaseClient, directory_rows: list[dict[str, Any]]) -> int:
    client.rest(
        "admin_user_directory?on_conflict=name",
        method="POST",
        payload=directory_rows,
        extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"},
    )
    return len(directory_rows)


def apply_directory_cleanup(client: SupabaseClient, row_ids: list[str]) -> int:
    if not row_ids:
        return 0
    client.rest(
        f"admin_user_directory?id=in.({','.join(row_ids)})",
        method="PATCH",
        payload={"is_active": False},
        extra_headers={"Prefer": "return=representation"},
    )
    return len(row_ids)


def apply_roles(client: SupabaseClient, role_dry_run: list[dict[str, str]]) -> int:
    payload = [
        {
            "user_id": row["user_id"],
            "role": row["next_role"],
        }
        for row in role_dry_run
        if row["current_role"] != row["next_role"]
    ]
    if not payload:
        return 0
    client.rest(
        "user_roles?on_conflict=user_id",
        method="POST",
        payload=payload,
        extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"},
    )
    return len(payload)


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync operational sites/workers with Supabase.")
    parser.add_argument("--site-xlsx", default=str(DEFAULT_SITE_XLSX))
    parser.add_argument("--workers-csv", default=str(DEFAULT_WORKERS_CSV))
    parser.add_argument("--email", default="admin.demo@inopnc.com")
    parser.add_argument("--password", default="1234qwer!")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--skip-sites", action="store_true")
    args = parser.parse_args()

    client = auth_client(args.email, args.password)
    profiles = fetch_rows(client, "profiles", "user_id,name,phone,affiliation,created_at,updated_at", "updated_at.desc")
    roles = fetch_rows(client, "user_roles", "user_id,role")
    try:
        directory = fetch_rows(
            client,
            "admin_user_directory",
            "id,name,phone,affiliation,notes,daily,source,is_active,linked_user_id,source_worker_id,role",
            "name.asc",
        )
    except HTTPError:
        directory = []
    try:
        pending_assignments = fetch_rows(
            client,
            "pending_role_assignments",
            "id,reserved_name,reserved_role,linked_user_id,status,note",
            "created_at.asc",
        )
    except HTTPError:
        pending_assignments = []
    try:
        sites = fetch_rows(client, "sites", "id,name,status,builder,company_name,source_dataset,created_at")
    except HTTPError:
        sites = fetch_rows(client, "sites", "id,name,status,created_at")

    site_rows = read_site_rows(Path(args.site_xlsx))
    worker_rows = read_worker_rows(Path(args.workers_csv))
    role_dry_run, blockers = build_role_plan(profiles, roles)
    directory_rows, worker_stats = build_directory_rows(worker_rows, profiles, roles)
    directory_rows = merge_pending_directory_rows(directory_rows, pending_assignments, directory)
    directory_plan = build_directory_plan(directory_rows, directory)
    site_plan = build_site_plan(site_rows, sites)

    result: dict[str, Any] = {
        "mode": "apply" if args.apply else "dry-run",
        "site_import_count": len(site_rows),
        "worker_import_count": len(worker_rows),
        "site_status_breakdown": dict(Counter(normalize_status(row.get("status")) for row in site_rows)),
        "site_plan": site_plan,
        "worker_match_plan": worker_stats,
        "directory_plan": {
            key: value
            for key, value in directory_plan.items()
            if key != "deactivate_ids"
        },
        "role_dry_run": role_dry_run,
        "blocking": blockers,
    }

    if args.apply:
        roles_applied = 0
        roles_skipped = 0
        role_apply_reason = "blocked_by_dry_run" if blockers else "applied"
        if blockers:
            roles_skipped = len([row for row in role_dry_run if row["current_role"] != row["next_role"]])
        else:
            roles_applied = apply_roles(client, role_dry_run)
        result["apply"] = {
            "sites_upserted": 0 if args.skip_sites else apply_sites(client, site_rows),
            "directory_upserted": apply_directory(client, directory_rows),
            "directory_deactivated": apply_directory_cleanup(client, directory_plan["deactivate_ids"]),
            "roles_applied": roles_applied,
            "roles_skipped": roles_skipped,
            "role_apply_reason": role_apply_reason,
        }

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
