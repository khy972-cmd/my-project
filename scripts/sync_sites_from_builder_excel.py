from __future__ import annotations

import argparse
import json
import re
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib import parse, request
from urllib.error import HTTPError

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"c:\Users\bobo\Desktop\site-manager-main_260302\INOPNC_FN-main\시공사-현장명 분리 요청 및 반영.xlsx")
STATUS_MAP = {
    "active": "진행중",
    "scheduled": "예정",
    "completed": "완료",
}
PROTECTED_SOURCES = {"manual"}
SAFE_UPDATE_SOURCES = {None, "", "site.xlsx", "test", "mock", "demo"}


def resolve_xlsx_path(raw_path: str | None) -> Path:
    if raw_path:
        candidate = Path(raw_path)
        if candidate.exists():
            return candidate

    if DEFAULT_XLSX.exists():
        return DEFAULT_XLSX

    workbook_dir = DEFAULT_XLSX.parent
    candidates = sorted(path for path in workbook_dir.glob("*.xlsx") if path.name != "site.xlsx")
    if len(candidates) == 1:
        return candidates[0]

    raise FileNotFoundError(f"workbook_not_found: {raw_path or str(DEFAULT_XLSX)}")


@dataclass
class SupabaseClient:
    base_url: str
    api_key: str
    access_token: str

    @property
    def headers(self) -> dict[str, str]:
        return {
            "apikey": self.api_key,
            "Authorization": f"Bearer {self.access_token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }

    def rest(
        self,
        path: str,
        method: str = "GET",
        payload: Any | None = None,
        extra_headers: dict[str, str] | None = None,
    ):
        url = f"{self.base_url}/rest/v1/{path}"
        headers = self.headers.copy()
        if extra_headers:
            headers.update(extra_headers)
        data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = request.Request(url, data=data, method=method, headers=headers)
        try:
            with request.urlopen(req) as response:
                body = response.read().decode("utf-8")
                return json.loads(body) if body else None
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"rest_error {method} {path}: {exc.code} {detail}") from exc


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    for line in (ROOT / ".env.local").read_text(encoding="utf-8").splitlines():
        if not line or line.strip().startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip()
    return env


def auth_client(email: str, password: str) -> SupabaseClient:
    env = load_env()
    base_url = env["VITE_SUPABASE_URL"].rstrip("/")
    api_key = env["VITE_SUPABASE_ANON_KEY"]
    req = request.Request(
        f"{base_url}/auth/v1/token?grant_type=password",
        data=json.dumps({"email": email, "password": password}).encode("utf-8"),
        headers={"apikey": api_key, "Content-Type": "application/json"},
        method="POST",
    )
    with request.urlopen(req) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return SupabaseClient(base_url=base_url, api_key=api_key, access_token=payload["access_token"])


def read_excel_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(value is not None and str(value).strip() != "" for value in row)
    ]


def load_legacy_mock_names() -> set[str]:
    text = (ROOT / "src" / "lib" / "siteList.ts").read_text(encoding="utf-8")
    return {
        match.group(1)
        for match in re.finditer(r'text:\s*"([^"]+)"', text)
    }


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_status(value: Any) -> str:
    raw = str(value or "").strip()
    return STATUS_MAP.get(raw.lower(), raw or "예정")


def serialize_created_at(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip()


def fetch_sites(client: SupabaseClient) -> list[dict[str, Any]]:
    query = parse.urlencode(
        {
            "select": "id,name,address,status,manager_name,manager_phone,created_at,updated_at,builder,company_name,source_dataset",
            "order": "created_at.asc,id.asc",
        },
        safe="(),*",
    )
    return client.rest(f"sites?{query}") or []


def build_import_rows(excel_rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], list[str]]:
    rows_by_id: dict[str, dict[str, Any]] = {}
    duplicate_ids: list[str] = []
    for row in excel_rows:
        site_id = str(row.get("id") or "").strip()
        if site_id in rows_by_id:
            duplicate_ids.append(site_id)
        rows_by_id[site_id] = {
            "id": site_id,
            "name": normalize_text(row.get("name")) or "",
            "address": None,
            "status": normalize_status(row.get("status")),
            "manager_name": None,
            "manager_phone": None,
            "builder": normalize_text(row.get("builder")),
            "company_name": normalize_text(row.get("company_name")),
            "created_at": serialize_created_at(row.get("created_at")),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "source_dataset": "site.xlsx",
        }
    return rows_by_id, duplicate_ids


def classify_delete_candidates(db_sites: list[dict[str, Any]], legacy_names: set[str]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for row in db_sites:
        site_id = str(row.get("id") or "").strip()
        name = str(row.get("name") or "").strip()
        source = row.get("source_dataset")

        if source == "manual":
            continue
        if source in {"test", "mock", "demo"}:
            candidates.append({"id": site_id, "name": name, "source_dataset": source, "reason": "source_dataset"})
            continue
        if site_id.lower().startswith("site") and site_id[4:].isdigit():
            candidates.append({"id": site_id, "name": name, "source_dataset": source, "reason": "legacy_static_id"})
            continue
        if name in legacy_names:
            candidates.append({"id": site_id, "name": name, "source_dataset": source, "reason": "legacy_static_name"})
            continue
        if name.startswith("CODEX ROLECHECK"):
            candidates.append({"id": site_id, "name": name, "source_dataset": source, "reason": "codex_seed_name"})
    return candidates


def build_plan(excel_rows: list[dict[str, Any]], db_sites: list[dict[str, Any]], legacy_names: set[str]) -> dict[str, Any]:
    import_rows, duplicate_ids = build_import_rows(excel_rows)
    existing_by_id = {str(row["id"]): row for row in db_sites}
    delete_candidates = classify_delete_candidates(db_sites, legacy_names)

    insert_count = 0
    update_count = 0
    conflicts: list[dict[str, Any]] = []
    for site_id in import_rows:
        existing = existing_by_id.get(site_id)
        if existing is None:
            insert_count += 1
            continue

        source = existing.get("source_dataset")
        if source in PROTECTED_SOURCES:
            conflicts.append({"id": site_id, "name": existing.get("name"), "source_dataset": source, "reason": "protected_manual_id_collision"})
        elif source not in SAFE_UPDATE_SOURCES:
            conflicts.append({"id": site_id, "name": existing.get("name"), "source_dataset": source, "reason": "protected_unknown_source_id_collision"})
        else:
            update_count += 1

    source_counts: dict[str, int] = {}
    for row in db_sites:
        key = str(row.get("source_dataset") or "(null)")
        source_counts[key] = source_counts.get(key, 0) + 1

    return {
        "excel_row_count": len(excel_rows),
        "duplicate_excel_id_count": len(duplicate_ids),
        "duplicate_excel_ids": duplicate_ids,
        "live_db_row_count": len(db_sites),
        "live_db_source_counts": source_counts,
        "delete_candidate_count": len(delete_candidates),
        "delete_candidates": delete_candidates,
        "import_insert_count": insert_count,
        "import_update_count": update_count,
        "import_conflict_count": len(conflicts),
        "import_conflicts": conflicts,
        "import_rows": list(import_rows.values()),
    }


def backup_candidates(delete_candidates: list[dict[str, Any]], db_sites: list[dict[str, Any]]) -> str | None:
    if not delete_candidates:
        return None

    by_id = {str(row["id"]): row for row in db_sites}
    backup_payload = [by_id[row["id"]] for row in delete_candidates if row["id"] in by_id]
    backup_path = Path(tempfile.gettempdir()) / f"inopnc_sites_backup_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.json"
    backup_path.write_text(json.dumps(backup_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(backup_path)


def delete_candidates_from_db(client: SupabaseClient, delete_candidates: list[dict[str, Any]]) -> int:
    if not delete_candidates:
        return 0
    encoded_ids = ",".join(row["id"] for row in delete_candidates)
    client.rest(f"sites?id=in.({encoded_ids})", method="DELETE", extra_headers={"Prefer": "return=representation"})
    return len(delete_candidates)


def upsert_sites(client: SupabaseClient, import_rows: list[dict[str, Any]]) -> int:
    client.rest(
        "sites?on_conflict=id",
        method="POST",
        payload=import_rows,
        extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"},
    )
    return len(import_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync builder/company site workbook into Supabase sites.")
    parser.add_argument("--xlsx")
    parser.add_argument("--email", default="admin.demo@inopnc.com")
    parser.add_argument("--password", default="1234qwer!")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    client = auth_client(args.email, args.password)
    xlsx_path = resolve_xlsx_path(args.xlsx)
    excel_rows = read_excel_rows(xlsx_path)
    legacy_names = load_legacy_mock_names()
    db_sites = fetch_sites(client)
    plan = build_plan(excel_rows, db_sites, legacy_names)

    result: dict[str, Any] = {
        "mode": "apply" if args.apply else "dry-run",
        "xlsx_path": str(xlsx_path),
        "live_db_row_count": plan["live_db_row_count"],
        "delete_candidate_count": plan["delete_candidate_count"],
        "delete_candidates": plan["delete_candidates"],
        "import_insert_count": plan["import_insert_count"],
        "import_update_count": plan["import_update_count"],
        "import_conflict_count": plan["import_conflict_count"],
        "import_conflicts": plan["import_conflicts"],
        "excel_row_count": plan["excel_row_count"],
        "duplicate_excel_id_count": plan["duplicate_excel_id_count"],
        "duplicate_excel_ids": plan["duplicate_excel_ids"],
        "live_db_source_counts": plan["live_db_source_counts"],
    }

    if args.apply:
        if plan["import_conflict_count"] > 0 or plan["duplicate_excel_id_count"] > 0:
            result["apply"] = {
                "applied": False,
                "reason": "blocked_by_conflict",
            }
        else:
            backup_path = backup_candidates(plan["delete_candidates"], db_sites)
            deleted_count = delete_candidates_from_db(client, plan["delete_candidates"])
            upserted_count = upsert_sites(client, plan["import_rows"])
            result["apply"] = {
                "applied": True,
                "backup_path": backup_path,
                "deleted_count": deleted_count,
                "upserted_count": upserted_count,
            }

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
